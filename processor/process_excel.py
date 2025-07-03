# Simple-Bot-for-Stock-and-Financial-Report-Data-Processing\processor\process_excel.py
from openpyxl import load_workbook
import pandas as pd
import os

from utils.excel_link_breaker import break_external_links
from utils.metric_inserter import add_new_metrics
from utils.formulas import compute_all_metrics
from config import new_metrics, new_profit_metrics, cashflow_metrics, percentage_metrics

def process_excel_file(file_path: str, output_path: str):
    cleaned_path = break_external_links(file_path)

    wb_formula = load_workbook(cleaned_path)
    wb_values = load_workbook(cleaned_path, data_only=True)
    ws_formula = wb_formula.active
    ws_values = wb_values.active

    # 加载各 sheet 数据
    df = pd.read_excel(cleaned_path, header=None).iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)
    df_profit = pd.read_excel(cleaned_path, sheet_name=1, header=None).iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)
    df_cashflow = pd.read_excel(cleaned_path, sheet_name=2, header=None).iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)
    #print("🔍 df.shape =", df.shape)
    #print("🔍 df_profit.shape =", df_profit.shape)
    #print("🔍 df_cashflow.shape =", df_cashflow.shape)


    # 插入行
    row_map = add_new_metrics(ws_formula, new_metrics, percentage_metrics)
    ws_profit = wb_formula.worksheets[1]
    profit_row_map = add_new_metrics(ws_profit, new_profit_metrics, percentage_metrics)
    ws_cashflow = wb_formula.worksheets[2]
    cashflow_row_map = add_new_metrics(ws_cashflow, cashflow_metrics, percentage_metrics)

    # 计算并写入
    all_metrics = compute_all_metrics(df, df_profit, df_cashflow)
    '''
    with open('all_metrics.txt', 'w', encoding='utf-8') as f:
        for metric, values in all_metrics.items():
            f.write(f'{metric}:\n')
            if isinstance(values, (list, pd.Series, pd.Index)):
                f.write(', '.join([str(v) for v in values]) + '\n\n')
            else:
                f.write(str(values) + '\n\n') 
    '''
    for sheet, mapping in zip(
        [ws_formula, ws_profit, ws_cashflow],
        [row_map, profit_row_map, cashflow_row_map]
    ):
        for metric, values in all_metrics.items():
            if metric not in mapping:
                continue
            row = mapping[metric]
            for i, v in enumerate(values):
                cell = sheet.cell(row=row, column=2 + i)
                if pd.notna(v):
                    cell.value = float(v)
                    cell.number_format = '0.0%' if metric in percentage_metrics else '0.00'
                else:
                    cell.value = None

    wb_formula.save(output_path)
    os.remove(cleaned_path)
    print(f"已处理并保存: {output_path}")
