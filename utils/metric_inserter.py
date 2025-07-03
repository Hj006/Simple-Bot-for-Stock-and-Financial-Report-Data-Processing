from openpyxl.styles import PatternFill
from copy import copy

def add_new_metrics(ws, new_metrics, percentage_metrics=None):
    """
    向指定 worksheet 中插入指标行，并返回 {指标名: 行号} 映射表。

    参数：
        ws: openpyxl 的 Worksheet 对象。
        new_metrics: List[str]，要插入的指标名称列表。
        percentage_metrics: 可选的百分比指标集合，用于标记格式。

    返回：
        dict: {指标名称: 插入后所在行号}
    """

    max_row = ws.max_row
    max_col = ws.max_column

    # 使用最后两行作为模板（白+灰）
    row_white = ws[max_row - 1]
    row_gray = ws[max_row]
    assert len(row_white) == len(row_gray)

    # 特殊插入逻辑配置
    insert_blank_after = {
        '扣非归母净利润',
        '货币资金/有息负债',
        '应付票据和账款（短期）',
        '应收变化/收入变化',
        '杠杆',
        '净利率',
        '归母扣非净利润 %（估算，不含投资收益）同比',
        '销售商品、提供劳务收到的现金+应收账款，应收票据增加额/营业收入',
        '经营活动产生的现金流量净额（同比）',
        '经营活动产生的现金流量净额（环比）',
    }

    insert_two_blank_before = {
        '每季度销售商品、提供劳务收到的现金（同比)',
        '每季度经营活动产生的现金流量净额（同比）'
    }

    green_fill_metrics = {
        '现金', '存货', '现金(环比）', '应收变化/收入变化',
        '每季度销售商品、提供劳务收到的现金（同比)',
        '每季度经营活动产生的现金流量净额（同比）'
    }

    yellow_fill_metrics = {'存货周转率', '应收账款周转率'}
    group_titles = {'同比', '环比'}

    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    yellow_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

    # 当前行位置从末尾继续添加
    current_row = max_row + 1

    # 插入两个空行
    for _ in range(2):
        for col in range(1, max_col + 1):
            ref_cell = row_white[col - 1]
            cell = ws.cell(row=current_row, column=col)
            cell.value = '' if col == 1 else None
            copy_style(cell, ref_cell)
        current_row += 1

    row_map = {}
    use_gray = False

    # 获取已有指标，避免重复插入
    # existing_metrics = {str(ws.cell(row=r, column=1).value).strip() for r in range(1, current_row)}
    '''
    
   existing_metrics = {
    str(ws.cell(row=r, column=1).value).strip()
    for r in range(1, current_row)
    if ws.cell(row=r, column=1).value is not None
}
 
    
    '''
    for metric in new_metrics:
        '''
        if metric in existing_metrics:
            continue  # 防止重复插入
        '''
        template_row = row_gray if use_gray else row_white
        use_gray = not use_gray

        # 插入标题类（同比/环比）
        if metric in group_titles:
            for col in range(1, max_col + 1):
                ref_cell = template_row[col - 1]
                cell = ws.cell(row=current_row, column=col)
                cell.value = metric if col == 1 else None
                copy_style(cell, ref_cell)
            current_row += 1
            continue

        if metric in insert_two_blank_before:
            for _ in range(2):
                for col in range(1, max_col + 1):
                    ref_cell = row_white[col - 1]
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = '' if col == 1 else None
                    copy_style(cell, ref_cell)
                current_row += 1



        # 插入正常指标
        for col in range(1, max_col + 1):
            ref_cell = template_row[col - 1]
            cell = ws.cell(row=current_row, column=col)
            cell.value = metric if col == 1 else None
            copy_style(cell, ref_cell)

            if col == 1:
                if metric in green_fill_metrics:
                    cell.fill = green_fill
                elif metric in yellow_fill_metrics:
                    cell.fill = yellow_fill

        row_map[metric] = current_row
        current_row += 1

        if metric in insert_blank_after:
            for col in range(1, max_col + 1):
                ref_cell = row_white[col - 1]
                cell = ws.cell(row=current_row, column=col)
                cell.value = '' if col == 1 else None
                copy_style(cell, ref_cell)
            current_row += 1

    return row_map

def copy_style(dest_cell, source_cell):
    dest_cell.font = copy(source_cell.font)
    dest_cell.alignment = copy(source_cell.alignment)
    dest_cell.border = copy(source_cell.border)
    dest_cell.fill = copy(source_cell.fill)
    dest_cell.number_format = copy(source_cell.number_format)
